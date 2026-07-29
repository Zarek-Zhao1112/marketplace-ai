#!/usr/bin/env python3
"""
库存同步脚本 - 用最新BSD库存数据更新所有历史JSON文件

使用方法：
    python sync_inventory.py <bsd_excel_path>

功能：
    1. 读取新的BSD Excel文件
    2. 提取库存数据（SKU -> 库存/履约方式/售价等）
    3. 更新所有历史JSON文件的库存相关字段
    4. 重新计算衍生指标（库存深度、处置建议等）
    5. 生成更新报告
"""

import sys
import os
import json
import glob
import math
from datetime import datetime
from openpyxl import load_workbook

# ── 配置 ──
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'sku_analysis')

# 需要从BSD更新的字段
BSD_FIELDS = [
    'Inventory',
    'FulfillmentType',
    'SellingPrice',
    'WarehouseLocation',
    'ItemCondition',
    'ActivationStatus',
    'SubcategoryName',
    'CategoryName',
    'ShortTitle',
]

# ── 库存深度计算 ──
def calc_inventory_depth(inventory):
    """库存深度层级"""
    try:
        inv = float(inventory)
    except (ValueError, TypeError):
        inv = 0
    if inv <= 0:
        return "零库存"
    elif inv <= 9:
        return "浅库存"
    elif inv <= 49:
        return "中库存"
    else:
        return "深库存"

# ── 处置建议计算 ──
def calc_disposal_suggestion(rma_pct, inventory, efficiency_level):
    """处置建议"""
    try:
        rma = abs(float(rma_pct))
    except (ValueError, TypeError):
        rma = 0
    try:
        inv = float(inventory)
    except (ValueError, TypeError):
        inv = 0
    
    # 风险等级
    if rma > 80:
        risk = "高危"
    elif rma >= 10:
        risk = "中危"
    else:
        risk = "低危"
    
    # 库存深度
    if inv <= 0:
        inv_depth = "零库存"
    elif inv <= 9:
        inv_depth = "浅库存"
    elif inv <= 49:
        inv_depth = "中库存"
    else:
        inv_depth = "深库存"
    
    # 处置建议
    if risk == "高危" and inv_depth in ["浅库存", "零库存"]:
        return "立即下架止损"
    elif risk == "高危" and inv_depth == "中库存":
        return "整改观察+限量销售，7天未改善则下架"
    elif risk == "高危" and inv_depth == "深库存":
        return "整改+清库存，7天观察期"
    elif risk == "中危" and efficiency_level in ["低动销", "零销负销"]:
        return "限制补货，优先清理库存"
    elif risk == "中危" and efficiency_level in ["核心主力", "潜力培育"]:
        return "维持销售，加强品质监控"
    elif risk == "低危" and efficiency_level == "零销负销":
        return "直接清退下架"
    elif risk == "低危" and efficiency_level == "低动销":
        return "评估是否保留"
    elif risk == "低危" and efficiency_level in ["核心主力", "潜力培育"]:
        return "正常运营，持续监控RMA变化"
    else:
        return "正常运营，持续监控RMA变化"

# ── 读取BSD Excel ──
def read_bsd_excel(bsd_path):
    """读取BSD Excel，提取库存数据"""
    print(f"读取BSD文件: {bsd_path}")
    
    wb = load_workbook(bsd_path, read_only=True)
    ws = wb.active
    
    # 找到表头行
    headers = {}
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                headers[str(val).strip()] = col
        if len(headers) > 5:
            break
    
    # 提取关键列
    sku_col = headers.get('NeweggItemNumber') or headers.get('Newegg Item Number')
    if not sku_col:
        print("错误: 找不到SKU列")
        return {}
    
    # 提取数据
    inventory_data = {}
    for row in range(row + 1, ws.max_row + 1):
        sku = ws.cell(row=row, column=sku_col).value
        if not sku:
            continue
        
        sku = str(sku).strip()
        record = {}
        
        for field in BSD_FIELDS:
            col = headers.get(field)
            if col:
                val = ws.cell(row=row, column=col).value
                record[field] = val
        
        if record:
            inventory_data[sku] = record
    
    wb.close()
    print(f"提取到 {len(inventory_data)} 个SKU的库存数据")
    return inventory_data

# ── 更新JSON文件 ──
def update_json_file(json_path, inventory_data):
    """更新单个JSON文件的库存数据"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    updated_count = 0
    records = data.get('records', [])
    
    for record in records:
        sku = record.get('NeweggItemNumber') or record.get('NeweggSku#')
        if not sku:
            continue
        
        sku = str(sku).strip()
        if sku in inventory_data:
            bsd_data = inventory_data[sku]
            
            # 更新库存相关字段
            for field in BSD_FIELDS:
                if field in bsd_data and bsd_data[field] is not None:
                    record[field] = bsd_data[field]
            
            # 更新衍生指标
            inventory = record.get('Inventory', 0)
            record['库存深度层级'] = calc_inventory_depth(inventory)
            
            # 更新处置建议
            rma_pct = record.get('RMA %', 0)
            efficiency = record.get('SKU效能等级', '低动销')
            record['处置建议'] = calc_disposal_suggestion(rma_pct, inventory, efficiency)
            
            updated_count += 1
    
    # 保存更新后的JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    return updated_count

# ── 主函数 ──
def main():
    if len(sys.argv) < 2:
        print("使用方法: python sync_inventory.py <bsd_excel_path>")
        sys.exit(1)
    
    bsd_path = sys.argv[1]
    if not os.path.exists(bsd_path):
        print(f"错误: 文件不存在 - {bsd_path}")
        sys.exit(1)
    
    # 1. 读取BSD数据
    inventory_data = read_bsd_excel(bsd_path)
    if not inventory_data:
        print("错误: 未提取到库存数据")
        sys.exit(1)
    
    # 2. 遍历所有JSON文件
    print(f"\n开始更新JSON文件...")
    print(f"数据目录: {DATA_DIR}")
    
    total_updated = 0
    files_updated = 0
    
    for seller_dir in os.listdir(DATA_DIR):
        seller_path = os.path.join(DATA_DIR, seller_dir)
        if not os.path.isdir(seller_path):
            continue
        
        for json_file in glob.glob(os.path.join(seller_path, '*.json')):
            try:
                count = update_json_file(json_file, inventory_data)
                if count > 0:
                    total_updated += count
                    files_updated += 1
                    print(f"  ✓ {seller_dir}/{os.path.basename(json_file)} - 更新{count}个SKU")
            except Exception as e:
                print(f"  ✗ {seller_dir}/{os.path.basename(json_file)} - 错误: {e}")
    
    # 3. 输出报告
    print(f"\n{'='*50}")
    print(f"库存同步完成")
    print(f"{'='*50}")
    print(f"BSD文件: {bsd_path}")
    print(f"更新文件数: {files_updated}")
    print(f"更新SKU数: {total_updated}")
    print(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

if __name__ == '__main__':
    main()
