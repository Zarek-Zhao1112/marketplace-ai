"""Excel导出工具 — 所有openpyxl格式化导出+图表逻辑"""
import io
import pandas as pd


def export_sku_multi_month(batches: list, seller_id: str) -> bytes:
    """导出多个月份的SKU数据到同一个Excel文件（每个sheet一个月份）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for batch in batches:
        date_readable = batch.get("date_readable", batch.get("date_period", "未知"))
        records = batch.get("records", [])
        if not records:
            continue

        df = pd.DataFrame(records)
        # 简化sheet名称（取日期范围的简短形式）
        sheet_name = date_readable.replace("/", "").replace(" ", "")[:31]  # Excel sheet名最长31字符
        ws = wb.create_sheet(title=sheet_name)

        # 写入表头
        headers = list(df.columns)
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 写入数据
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else "")
                cell.font = normal_font
                cell.border = thin_border

        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sku_excel(df: pd.DataFrame) -> bytes:
    """导出SKU分级治理总表，首页为数值说明sheet"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    # ── 样式 ──
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11, color="4472C4")
    normal_font = Font(size=10)
    note_font = Font(size=9, italic=True, color="666666")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_row(ws, row, data, font=normal_font, fill=None):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = font
            cell.border = thin_border
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ══════════════════════════════════════════════════════
    #  Sheet 1: 数值说明
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "数值说明"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 55

    row = 1
    ws.cell(row=row, column=1, value="SKU分级治理总表 — 数值来源与计算说明").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="本页说明总表中各字段的数据来源、计算公式和分级规则").font = note_font
    row += 2

    # ── 一、原始字段（直接取自销售表） ──
    ws.cell(row=row, column=1, value="一、原始字段（直接取自销售表/库存表）").font = section_font
    row += 1
    write_row(ws, row, ("字段名", "数据来源", "说明"), font=header_font, fill=header_fill)
    row += 1
    raw_fields = [
        ("Item Description", "销售表第1列", "商品描述，含SKU编号、品牌、型号等信息"),
        ("GMV", "销售表 GMV 列", "该SKU在统计周期内的总销售额（美元）"),
        ("RMA %", "销售表 RMA% 列", "退货率，负值表示有退货，如 -0.93% 表示退货金额占GMV的0.93%"),
        ("Total Margin (without EIMS)", "销售表 Margin 列", "该SKU的总毛利（扣除EIMS前）"),
        ("Net Quantity Sold", "销售表 Quantity 列", "该SKU在统计周期内的净销量（件）"),
        ("SKU Count", "销售表", "SKU数量，通常为1（单品）"),
        ("NeweggItemNumber", "销售表", "Newegg商品编号，用于匹配库存表"),
        ("SubcategoryName", "销售表", "商品子品类（如 Video Card - Nvidia）"),
        ("ActivationStatus", "销售表", "上架状态：Active / Inactive"),
        ("FulfillmentType", "销售表", "履约方式：Ship by Newegg / Ship by Seller"),
        ("WarehouseLocation", "销售表", "仓库位置（如 United States - 09, Hong Kong）"),
        ("Inventory", "库存表（按NeweggItemNumber匹配）", "当前库存数量（件）"),
        ("SellingPrice", "销售表", "商品售价（美元）"),
        ("ItemCondition", "销售表", "商品成色：New / Used - Like New / Used - Very Good / Refurbished 等"),
    ]
    for f in raw_fields:
        write_row(ws, row, f)
        row += 1
    row += 1

    # ── 二、衍生字段（系统计算） ──
    ws.cell(row=row, column=1, value="二、衍生字段（系统自动计算）").font = section_font
    row += 1
    write_row(ws, row, ("字段名", "计算公式", "说明"), font=header_font, fill=header_fill)
    row += 1
    derived_fields = [
        ("客单价", "= GMV ÷ Net Quantity Sold", "每件商品的平均销售额，非直接售价"),
        ("单件毛利", "= Total Margin ÷ Net Quantity Sold", "每件商品的平均毛利"),
        ("单SKU毛利率(%)", "= Total Margin ÷ GMV × 100%", "该SKU的毛利率百分比"),
        ("GMV贡献占比(%)", "= 该SKU的GMV ÷ 全店所有SKU的GMV总和 × 100%", "该SKU在全店GMV中的占比"),
        ("毛利贡献占比(%)", "= 该SKU的Margin ÷ 全店所有SKU的Margin总和 × 100%", "该SKU在全店毛利中的占比"),
        ("退货损失金额", "= |GMV| × |RMA%| ÷ 100", "退货造成的金额损失，如 RMA%=-0.93% 则损失=GMV×0.0093"),
        ("退货件数", "= ceil(|RMA%|/100 × Net Quantity Sold ÷ (1 - |RMA%|/100))", "从RMA%和销量反推退货件数（向上取整，RMA%≤0或≥100%时返回0）"),
        ("退货毛利侵蚀率", "= 退货损失金额 ÷ Total Margin", "退货损失占毛利的比例，越低越好"),
        ("日均销量", "= Net Quantity Sold ÷ 实际天数", "日均销量（按实际数据周期天数计算）"),
    ]
    for f in derived_fields:
        write_row(ws, row, f)
        row += 1
    row += 1

    # ── 三、分级标签（规则判定） ──
    ws.cell(row=row, column=1, value="三、分级标签（按规则自动判定）").font = section_font
    row += 1
    write_row(ws, row, ("字段名", "分级规则", "说明"), font=header_font, fill=header_fill)
    row += 1
    level_fields = [
        ("库存深度层级",
         "零库存: ≤0件\n浅库存: 1-9件\n中库存: 10-49件\n深库存: ≥50件",
         "按当前库存数量划分"),
        ("SKU风险等级",
         "低危: |RMA%| < 10%\n中危: 10% ≤ |RMA%| < 80%\n高危: |RMA%| ≥ 80%",
         "按退货率绝对值划分，退货率越高风险越大"),
        ("SKU效能等级",
         "零销负销: 销量 = 0\n低动销: 销量 1-2件\n潜力培育: 销量 3-9件\n核心主力: 销量 ≥10件",
         "按统计周期内总销量划分"),
        ("客单价分层",
         "低客单: ≤$100\n中客单: $100-$500\n高客单: >$500",
         "按客单价（GMV÷销量）划分"),
        ("整改优先级得分",
         "= 退货损失(40%) + 毛利侵蚀(30%) + RMA严重度(20%) + 动销逆向(10%)",
         "退货损失: min(100, 损失/$100×100); 毛利侵蚀: min(100, 侵蚀率×100); RMA: min(100, |RMA%|); 动销: 0件=50分, 1-2件=80分, 3-5件=50分, 6-10件=20分, >10件=0分"),
        ("整改优先级",
         "极高: ≥40分\n高: ≥25分\n中: ≥10分\n低: <10分",
         "按绝对分数分级（不再使用百分位排名）"),
        ("处置建议",
         "由 风险等级 × 库存深度 × 效能等级 组合判定",
         "系统自动生成的处置建议（见下方矩阵）"),
    ]
    for f in level_fields:
        write_row(ws, row, f)
        row += 1
    row += 1

    # ── 四、处置建议矩阵 ──
    ws.cell(row=row, column=1, value="四、处置建议判定矩阵").font = section_font
    row += 1
    write_row(ws, row, ("风险等级", "库存深度", "效能等级", "处置建议"), font=header_font, fill=header_fill)
    row += 1
    matrix = [
        ("高危", "浅/零库存", "任意", "立即下架止损"),
        ("高危", "中库存", "任意", "整改观察+限量销售，7天未改善则下架"),
        ("高危", "深库存", "任意", "整改+清库存，7天观察期，同步启动退货流程"),
        ("中危", "任意", "低动销/零销负销", "限制补货，优先清理库存，观察30天"),
        ("中危", "任意", "核心主力/潜力培育", "维持现有销售，加强品质监控，月度复查"),
        ("低危", "任意", "零销负销", "直接清退下架"),
        ("低危", "任意", "低动销", "评估是否保留，无战略价值建议清退"),
        ("低危", "任意", "核心主力/潜力培育", "正常运营，持续监控RMA变化"),
    ]
    for m in matrix:
        write_row(ws, row, m)
        row += 1
    row += 1

    # ── 五、商品成色/品类/品牌提取规则 ──
    ws.cell(row=row, column=1, value="五、商品信息提取规则").font = section_font
    row += 1
    write_row(ws, row, ("字段", "提取逻辑", "说明"), font=header_font, fill=header_fill)
    row += 1
    extract_fields = [
        ("商品成色", "从 Item Description 中匹配关键词", "New=全新, Used/Like New/Very Good=翻新, Refurbished=翻新"),
        ("品类", "从 Item Description 中匹配品类关键词", "显卡/主板/内存/SSD/CPU/风扇/鼠标/工具等"),
        ("品牌", "从 Item Description 中提取品牌名", "MSI/Sapphire/Intel/CornE/Noctua/LIAN LI 等"),
    ]
    for f in extract_fields:
        write_row(ws, row, f)
        row += 1

    # ══════════════════════════════════════════════════════
    #  Sheet 2: 数据（原始DataFrame）
    # ══════════════════════════════════════════════════════
    ws_data = wb.create_sheet("SKU销售数据表", 1)
    headers = list(df.columns)
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row_data, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=val).font = normal_font

    # 自动列宽
    for col in ws_data.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_seller_history_excel(hist_df: pd.DataFrame, seller_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()

    # ── 计算说明 Sheet ──
    ws_guide = wb.active
    ws_guide.title = "计算说明"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=12)
    normal_font = Font(size=10)

    guide_content = [
        ("指标", "满分", "计算逻辑"),
        ("GMV（销售额）", "30分", "线性比例：GMV ÷ $50,000 × 30，上限30分"),
        ("总毛利", "25分", "线性比例：毛利 ÷ $10,000 × 25，上限25分"),
        ("RMA%（退货率）", "20分", "≤2%得20分，≤5%得16分，≤8%得12分，≤15%得8分，≤25%得4分"),
        ("总销量", "10分", "≥50件10分，≥20件7分，≥5件4分，>0件2分"),
        ("SKU数", "10分", "≥20个10分，≥10个7分，≥5个4分，>0个2分"),
        ("毛利率", "5分", "≥10%得5分，≥5%得4分，>0%得2分"),
        ("合计", "100分", "满分100分"),
    ]

    for row_idx, (col1, col2, col3) in enumerate(guide_content, 1):
        ws_guide.cell(row=row_idx, column=1, value=col1)
        ws_guide.cell(row=row_idx, column=2, value=col2)
        ws_guide.cell(row=row_idx, column=3, value=col3)
        if row_idx == 1:
            for col in range(1, 4):
                ws_guide.cell(row=row_idx, column=col).fill = header_fill
                ws_guide.cell(row=row_idx, column=col).font = header_font
                ws_guide.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
        else:
            for col in range(1, 4):
                ws_guide.cell(row=row_idx, column=col).font = normal_font

    ws_guide.column_dimensions["A"].width = 18
    ws_guide.column_dimensions["B"].width = 10
    ws_guide.column_dimensions["C"].width = 55

    # 等级说明
    grade_start_row = len(guide_content) + 2
    ws_guide.cell(row=grade_start_row, column=1, value="等级划分").font = title_font
    grade_content = [
        ("等级", "分数区间", "标签", "颜色"),
        ("A", "≥75分", "核心优质卖家", "#52c41a（绿色）"),
        ("B", "≥60分", "高潜力卖家", "#1890ff（蓝色）"),
        ("C", "≥45分", "普通合规卖家", "#faad14（黄色）"),
        ("D", "<45分", "高风险卖家", "#f5222d（红色）"),
    ]
    for row_idx, (g, score, label, color) in enumerate(grade_content, grade_start_row + 1):
        ws_guide.cell(row=row_idx, column=1, value=g)
        ws_guide.cell(row=row_idx, column=2, value=score)
        ws_guide.cell(row=row_idx, column=3, value=label)
        ws_guide.cell(row=row_idx, column=4, value=color)
        if row_idx == grade_start_row + 1:
            for col in range(1, 5):
                ws_guide.cell(row=row_idx, column=col).fill = header_fill
                ws_guide.cell(row=row_idx, column=col).font = header_font

    # 行业基准说明
    bench_start_row = grade_start_row + len(grade_content) + 2
    ws_guide.cell(row=bench_start_row, column=1, value="行业基准（3C/消费电子）").font = title_font
    bench_content = [
        ("指标", "优秀", "良好", "一般", "较差"),
        ("RMA%", "<3%", "3-5%", "5-8%", ">8%"),
        ("毛利率", ">15%", "10-15%", "5-10%", "<5%"),
        ("月GMV", ">$50K", "$20-50K", "$5-20K", "<$5K"),
        ("SKU动销率", ">80%", "60-80%", "40-60%", "<40%"),
        ("库存周转", ">8次/年", "5-8次", "3-5次", "<3次"),
    ]
    for row_idx, row_data in enumerate(bench_content, bench_start_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            ws_guide.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == bench_start_row + 1:
                ws_guide.cell(row=row_idx, column=col_idx).fill = header_fill
                ws_guide.cell(row=row_idx, column=col_idx).font = header_font

    ws_guide.column_dimensions["D"].width = 12
    ws_guide.column_dimensions["E"].width = 12

    # 备注
    note_row = bench_start_row + len(bench_content) + 2
    ws_guide.cell(row=note_row, column=1, value="备注").font = title_font
    ws_guide.cell(row=note_row + 1, column=1, value="1. 卖家数量<10时，行业基准权重60%，内部数据权重40%").font = normal_font
    ws_guide.cell(row=note_row + 2, column=1, value="2. 卖家数量10-20时，行业基准权重30%，内部数据权重70%").font = normal_font
    ws_guide.cell(row=note_row + 3, column=1, value="3. 卖家数量≥20时，行业基准权重10%，内部数据权重90%").font = normal_font

    # ── 卖家健康度数据 Sheet ──
    ws_data = wb.create_sheet("卖家健康度数据", 0)  # 插入到第一个位置
    headers = list(hist_df.columns)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(hist_df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=val)
    for col in ws_data.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    if len(hist_df) >= 2 and "健康度评分" in hist_df.columns:
        ws_chart = wb.create_sheet("健康度趋势")
        score_col = headers.index("健康度评分") + 1
        date_col = 1
        data_ref = Reference(ws_data, min_col=score_col, min_row=1, max_row=len(hist_df) + 1)
        cats_ref = Reference(ws_data, min_col=date_col, min_row=2, max_row=len(hist_df) + 1)
        chart = LineChart()
        chart.title = f"卖家 {seller_id} 健康度评分趋势"
        chart.y_axis.title = "评分"
        chart.x_axis.title = "日期"
        chart.style = 10
        chart.width = 24
        chart.height = 14
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.line.width = 28000
        ws_chart.add_chart(chart, "A1")
    elif len(hist_df) >= 2:
        ws_chart = wb.create_sheet("健康度趋势")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sku_history_excel(batches: list, seller_id: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    all_records = []
    for batch in batches:
        all_records.extend(batch.get("records", []))

    if not all_records:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    df = pd.DataFrame(all_records)
    ws_data = wb.active
    ws_data.title = "SKU明细数据"
    headers = list(df.columns)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws_data.cell(row=row_idx, column=col_idx, value=val if pd.notna(val) else "")
    for col in ws_data.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_data.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    def _write_pivot_sheet(wb, sheet_name, key_col, value_col, chart_title, chart_type="bar", agg_func="sum"):
        if key_col not in df.columns:
            return
        if agg_func == "count":
            pivot = df.groupby(key_col, dropna=False).size().reset_index(name="数值")
            pivot = pivot.rename(columns={key_col: "类别"})
        else:
            pivot = df.groupby(key_col, dropna=False).agg({value_col: "sum"}).reset_index()
            pivot = pivot.rename(columns={key_col: "类别", value_col: "数值"})
        pivot = pivot.sort_values("数值", ascending=False)

        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value="类别").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="数值").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        for r_idx, (_, row) in enumerate(pivot.iterrows(), 2):
            ws.cell(row=r_idx, column=1, value=row["类别"])
            ws.cell(row=r_idx, column=2, value=row["数值"])
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(pivot) + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(pivot) + 1)

        if chart_type == "pie":
            chart = PieChart()
            chart.style = 10
        else:
            chart = BarChart()
            chart.style = 10
            chart.y_axis.title = "数量"
        chart.title = chart_title
        chart.width = 22
        chart.height = 14
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "D1")

    risk_col = next((c for c in ["SKU风险等级", "风险等级"] if c in df.columns), None)
    if risk_col:
        _write_pivot_sheet(wb, "风险等级分布", risk_col, risk_col, "风险等级分布", "pie", "count")

    category_col = next((c for c in ["品类", "品类名称", "Category"] if c in df.columns), None)
    if category_col:
        gmv_col = next((c for c in ["GMV", "gmv", "GMV($)"] if c in df.columns), None)
        if gmv_col:
            _write_pivot_sheet(wb, "品类销售分布", category_col, gmv_col, "品类GMV分布")
        else:
            _write_pivot_sheet(wb, "品类销售分布", category_col, category_col, "品类SKU分布", "count")

    inv_col = next((c for c in ["库存深度层级", "库存深度", "Inventory Depth"] if c in df.columns), None)
    if inv_col:
        _write_pivot_sheet(wb, "库存深度分布", inv_col, inv_col, "库存深度分布", "bar", "count")

    eff_col = next((c for c in ["SKU效能等级", "效能等级", "Efficiency"] if c in df.columns), None)
    if eff_col:
        _write_pivot_sheet(wb, "效能等级分布", eff_col, eff_col, "效能等级分布", "bar", "count")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_all_sellers_pano(all_history: dict) -> tuple:
    """导出所有卖家全景报表（4个Sheet：计算说明、卖家全景概览、指标分布分析、卖家风险矩阵）"""
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, BarChart, ScatterChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import math

    # 提取日期范围
    all_dates = set()
    for hist in all_history.values():
        for rec in hist:
            date_str = rec.get("日期", "")
            if date_str:
                all_dates.add(date_str)

    if all_dates:
        # 解析日期范围，取最早开始日和最晚结束日
        start_dates = []
        end_dates = []
        for d in all_dates:
            parts = d.split(" - ")
            if len(parts) == 2:
                start_dates.append(parts[0].replace("/", ""))
                end_dates.append(parts[1].replace("/", ""))
        if start_dates and end_dates:
            date_range = f"{min(start_dates)}-{max(end_dates)}"
        else:
            date_range = list(all_dates)[0].replace("/", "").replace(" ", "")
    else:
        date_range = "unknown"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=12)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    grade_colors = {"A": "52c41a", "B": "1890ff", "C": "faad14", "D": "f5222d"}
    grade_labels = {"A": "核心优质", "B": "高潜力", "C": "普通合规", "D": "高风险"}

    def write_header(ws, headers, row=1):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_col_width(ws):
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    # ── 收集所有卖家最新数据 ──
    overview_data = []
    for sid, hist in all_history.items():
        if not hist:
            continue
        latest = hist[-1]
        gmv = latest.get("GMV", 0) or 0
        rma = latest.get("RMA%", 0) or 0
        margin = latest.get("总毛利", 0) or 0
        overview_data.append({
            "卖家ID": sid,
            "健康度评分": latest.get("健康度评分", 0),
            "等级": latest.get("等级", "D"),
            "GMV": gmv,
            "RMA%": abs(rma),
            "总毛利": margin,
            "毛利率": round(margin / gmv * 100, 2) if gmv > 0 else 0,
            "总销量": latest.get("总销量", 0) or 0,
            "SKU数": latest.get("SKU数", 0) or 0,
        })

    total_sellers = len(overview_data)
    total_gmv = sum(d["GMV"] for d in overview_data)

    wb = Workbook()

    # ══════════════════════════════════════════════════════
    #  Sheet 1: 计算说明
    # ══════════════════════════════════════════════════════
    ws_guide = wb.active
    ws_guide.title = "1.计算说明"
    guide = [
        ("指标", "满分", "计算逻辑"),
        ("GMV（销售额）", "30分", "线性比例：GMV ÷ $50,000 × 30，上限30分"),
        ("总毛利", "25分", "线性比例：毛利 ÷ $10,000 × 25，上限25分"),
        ("RMA%（退货率）", "20分", "≤2%得20分，≤5%得16分，≤8%得12分，≤15%得8分，≤25%得4分"),
        ("总销量", "10分", "≥50件10分，≥20件7分，≥5件4分，>0件2分"),
        ("SKU数", "10分", "≥20个10分，≥10个7分，≥5个4分，>0个2分"),
        ("毛利率", "5分", "≥10%得5分，≥5%得4分，>0%得2分"),
        ("合计", "100分", ""),
    ]
    write_header(ws_guide, guide[0], 1)
    for i, row in enumerate(guide[1:], 2):
        for j, val in enumerate(row, 1):
            ws_guide.cell(row=i, column=j, value=val).font = normal_font

    grade_row = len(guide) + 2
    ws_guide.cell(row=grade_row, column=1, value="【等级划分】").font = title_font
    write_header(ws_guide, ("等级", "分数区间", "标签"), grade_row + 1)
    for i, (g, score, label) in enumerate([("A", "≥75分", "核心优质"), ("B", "≥60分", "高潜力"), ("C", "≥45分", "普通合规"), ("D", "<45分", "高风险")], grade_row + 2):
        for j, val in enumerate([g, score, label], 1):
            ws_guide.cell(row=i, column=j, value=val).font = normal_font

    bench_row = grade_row + 6
    ws_guide.cell(row=bench_row, column=1, value="【行业基准（3C/消费电子）】").font = title_font
    write_header(ws_guide, ("指标", "优秀", "良好", "一般", "较差"), bench_row + 1)
    for i, row in enumerate([("RMA%", "<3%", "3-5%", "5-8%", ">8%"), ("毛利率", ">15%", "10-15%", "5-10%", "<5%"), ("月GMV", ">$50K", "$20-50K", "$5-20K", "<$5K")], bench_row + 2):
        for j, val in enumerate(row, 1):
            ws_guide.cell(row=i, column=j, value=val).font = normal_font
    auto_col_width(ws_guide)
    ws_guide.column_dimensions["C"].width = 55

    # ══════════════════════════════════════════════════════
    #  Sheet 2: 卖家全景概览（饼图 + GMV贡献 + Top10低分）
    # ══════════════════════════════════════════════════════
    ws_overview = wb.create_sheet("2.卖家全景概览", 1)

    # 统计等级分布
    grade_counts = {"A": 0, "B": 0, "C": 0, "D": 0}
    grade_gmv = {"A": 0, "B": 0, "C": 0, "D": 0}
    for d in overview_data:
        g = d["等级"]
        grade_counts[g] = grade_counts.get(g, 0) + 1
        grade_gmv[g] = grade_gmv.get(g, 0) + d["GMV"]

    # 写入概览数据表
    overview_headers = ["卖家ID", "健康度评分", "等级", "GMV", "RMA%", "毛利率", "总销量", "SKU数"]
    write_header(ws_overview, overview_headers)
    sorted_data = sorted(overview_data, key=lambda x: x["健康度评分"], reverse=True)
    for row_idx, d in enumerate(sorted_data, 2):
        for col_idx, key in enumerate(overview_headers, 1):
            cell = ws_overview.cell(row=row_idx, column=col_idx, value=d[key])
            cell.font = normal_font
            cell.border = thin_border
            if key == "等级" and d[key] in grade_colors:
                cell.fill = PatternFill(start_color=grade_colors[d[key]], end_color=grade_colors[d[key]], fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)
    auto_col_width(ws_overview)

    # 等级分布饼图（数据写在单独的Sheet）
    ws_pie_data = wb.create_sheet("2a.等级分布数据")
    ws_pie_data.cell(row=1, column=1, value="等级").font = header_font
    ws_pie_data.cell(row=1, column=1).fill = header_fill
    ws_pie_data.cell(row=1, column=2, value="数量").font = header_font
    ws_pie_data.cell(row=1, column=2).fill = header_fill
    ws_pie_data.cell(row=1, column=3, value="占比").font = header_font
    ws_pie_data.cell(row=1, column=3).fill = header_fill
    for i, g in enumerate(["A", "B", "C", "D"], 2):
        ws_pie_data.cell(row=i, column=1, value=f"{g}-{grade_labels[g]}").font = normal_font
        ws_pie_data.cell(row=i, column=2, value=grade_counts[g]).font = normal_font
        pct = round(grade_counts[g] / total_sellers * 100, 1) if total_sellers > 0 else 0
        ws_pie_data.cell(row=i, column=3, value=f"{pct}%").font = normal_font
    auto_col_width(ws_pie_data)

    pie = PieChart()
    pie.title = f"卖家等级分布（共{total_sellers}个）"
    pie.style = 10
    pie.add_data(Reference(ws_pie_data, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    pie.set_categories(Reference(ws_pie_data, min_col=1, min_row=2, max_row=5))
    pie.width = 18
    pie.height = 12
    ws_pie_data.add_chart(pie, "E1")

    # GMV贡献柱状图（数据写在单独的Sheet）
    ws_gmv_data = wb.create_sheet("2b.GMV贡献数据")
    ws_gmv_data.cell(row=1, column=1, value="等级").font = header_font
    ws_gmv_data.cell(row=1, column=1).fill = header_fill
    ws_gmv_data.cell(row=1, column=2, value="GMV总额").font = header_font
    ws_gmv_data.cell(row=1, column=2).fill = header_fill
    ws_gmv_data.cell(row=1, column=3, value="GMV占比").font = header_font
    ws_gmv_data.cell(row=1, column=3).fill = header_fill
    for i, g in enumerate(["A", "B", "C", "D"], 2):
        ws_gmv_data.cell(row=i, column=1, value=f"{g}-{grade_labels[g]}").font = normal_font
        ws_gmv_data.cell(row=i, column=2, value=round(grade_gmv[g], 2)).font = normal_font
        pct = round(grade_gmv[g] / total_gmv * 100, 1) if total_gmv > 0 else 0
        ws_gmv_data.cell(row=i, column=3, value=f"{pct}%").font = normal_font
    auto_col_width(ws_gmv_data)

    bar = BarChart()
    bar.type = "col"
    bar.title = "各等级卖家GMV贡献"
    bar.y_axis.title = "GMV ($)"
    bar.style = 10
    bar.add_data(Reference(ws_gmv_data, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    bar.set_categories(Reference(ws_gmv_data, min_col=1, min_row=2, max_row=5))
    bar.width = 18
    bar.height = 12
    ws_gmv_data.add_chart(bar, "E1")

    # Top10低分卖家
    ws_top10 = wb.create_sheet("2c.Top10低分卖家")
    write_header(ws_top10, ["排名", "卖家ID", "健康度评分", "等级", "GMV", "RMA%", "毛利率", "风险提示"])
    low_sellers = sorted(overview_data, key=lambda x: x["健康度评分"])[:10]
    risk_tips = {
        "D": "高风险卖家，建议优化或清退",
        "C": "普通合规，需持续监控",
        "B": "高潜力，保持关注",
        "A": "核心优质，维持现状",
    }
    for i, d in enumerate(low_sellers, 1):
        ws_top10.cell(row=i + 1, column=1, value=i).font = normal_font
        ws_top10.cell(row=i + 1, column=2, value=d["卖家ID"]).font = normal_font
        ws_top10.cell(row=i + 1, column=3, value=d["健康度评分"]).font = normal_font
        cell_grade = ws_top10.cell(row=i + 1, column=4, value=d["等级"])
        cell_grade.font = Font(color="FFFFFF", bold=True, size=10)
        if d["等级"] in grade_colors:
            cell_grade.fill = PatternFill(start_color=grade_colors[d["等级"]], end_color=grade_colors[d["等级"]], fill_type="solid")
        ws_top10.cell(row=i + 1, column=5, value=round(d["GMV"], 2)).font = normal_font
        ws_top10.cell(row=i + 1, column=6, value=d["RMA%"]).font = normal_font
        ws_top10.cell(row=i + 1, column=7, value=d["毛利率"]).font = normal_font
        ws_top10.cell(row=i + 1, column=8, value=risk_tips.get(d["等级"], "")).font = normal_font
    auto_col_width(ws_top10)
    ws_top10.column_dimensions["H"].width = 35

    # ══════════════════════════════════════════════════════
    #  Sheet 3: 指标分布分析（GMV/RMA%/毛利率分布）
    # ══════════════════════════════════════════════════════
    ws_dist = wb.create_sheet("3.指标分布分析", 2)

    def calc_distribution(values, bins):
        """计算分布区间"""
        counts = [0] * len(bins)
        for v in values:
            for i, (low, high) in enumerate(bins):
                if low <= v < high:
                    counts[i] += 1
                    break
        return counts

    # GMV分布
    gmv_bins = [("<$5K", 0, 5000), ("$5K-20K", 5000, 20000), ("$20K-50K", 20000, 50000), (">$50K", 50000, float('inf'))]
    gmv_values = [d["GMV"] for d in overview_data]
    ws_dist.cell(row=1, column=1, value="【GMV分布】").font = title_font
    write_header(ws_dist, ["区间", "卖家数量", "占比"], 2)
    for i, (label, low, high) in enumerate(gmv_bins, 3):
        count = sum(1 for v in gmv_values if low <= v < high)
        ws_dist.cell(row=i, column=1, value=label).font = normal_font
        ws_dist.cell(row=i, column=2, value=count).font = normal_font
        pct = round(count / total_sellers * 100, 1) if total_sellers > 0 else 0
        ws_dist.cell(row=i, column=3, value=f"{pct}%").font = normal_font

    # RMA%分布
    rma_bins = [("<3%", 0, 3), ("3-5%", 3, 5), ("5-8%", 5, 8), ("8-15%", 8, 15), (">15%", 15, float('inf'))]
    rma_values = [d["RMA%"] for d in overview_data]
    ws_dist.cell(row=8, column=1, value="【RMA%分布】").font = title_font
    write_header(ws_dist, ["区间", "卖家数量", "占比"], 9)
    for i, (label, low, high) in enumerate(rma_bins, 10):
        count = sum(1 for v in rma_values if low <= v < high)
        ws_dist.cell(row=i, column=1, value=label).font = normal_font
        ws_dist.cell(row=i, column=2, value=count).font = normal_font
        pct = round(count / total_sellers * 100, 1) if total_sellers > 0 else 0
        ws_dist.cell(row=i, column=3, value=f"{pct}%").font = normal_font

    # 毛利率分布
    margin_bins = [("<5%", 0, 5), ("5-10%", 5, 10), ("10-15%", 10, 15), ("15-20%", 15, 20), (">20%", 20, float('inf'))]
    margin_values = [d["毛利率"] for d in overview_data]
    ws_dist.cell(row=16, column=1, value="【毛利率分布】").font = title_font
    write_header(ws_dist, ["区间", "卖家数量", "占比"], 17)
    for i, (label, low, high) in enumerate(margin_bins, 18):
        count = sum(1 for v in margin_values if low <= v < high)
        ws_dist.cell(row=i, column=1, value=label).font = normal_font
        ws_dist.cell(row=i, column=2, value=count).font = normal_font
        pct = round(count / total_sellers * 100, 1) if total_sellers > 0 else 0
        ws_dist.cell(row=i, column=3, value=f"{pct}%").font = normal_font
    auto_col_width(ws_dist)

    # GMV分布柱状图
    gmv_chart = BarChart()
    gmv_chart.type = "col"
    gmv_chart.title = "卖家GMV分布"
    gmv_chart.y_axis.title = "卖家数量"
    gmv_chart.style = 10
    gmv_chart.add_data(Reference(ws_dist, min_col=2, min_row=2, max_row=6), titles_from_data=True)
    gmv_chart.set_categories(Reference(ws_dist, min_col=1, min_row=3, max_row=6))
    gmv_chart.width = 16
    gmv_chart.height = 10
    ws_dist.add_chart(gmv_chart, "E2")

    # RMA%分布柱状图
    rma_chart = BarChart()
    rma_chart.type = "col"
    rma_chart.title = "卖家RMA%分布"
    rma_chart.y_axis.title = "卖家数量"
    rma_chart.style = 10
    rma_chart.add_data(Reference(ws_dist, min_col=2, min_row=9, max_row=14), titles_from_data=True)
    rma_chart.set_categories(Reference(ws_dist, min_col=1, min_row=10, max_row=14))
    rma_chart.width = 16
    rma_chart.height = 10
    ws_dist.add_chart(rma_chart, "E17")

    # 毛利率分布柱状图
    margin_chart = BarChart()
    margin_chart.type = "col"
    margin_chart.title = "卖家毛利率分布"
    margin_chart.y_axis.title = "卖家数量"
    margin_chart.style = 10
    margin_chart.add_data(Reference(ws_dist, min_col=2, min_row=17, max_row=22), titles_from_data=True)
    margin_chart.set_categories(Reference(ws_dist, min_col=1, min_row=18, max_row=22))
    margin_chart.width = 16
    margin_chart.height = 10
    ws_dist.add_chart(margin_chart, "E32")

    # ══════════════════════════════════════════════════════
    #  Sheet 4: 卖家风险矩阵（GMV vs RMA% 散点图）
    # ══════════════════════════════════════════════════════
    ws_risk = wb.create_sheet("4.卖家风险矩阵", 3)

    # 写入散点图数据
    write_header(ws_risk, ["卖家ID", "GMV", "RMA%", "等级", "健康度评分"])
    for i, d in enumerate(sorted_data, 2):
        ws_risk.cell(row=i, column=1, value=d["卖家ID"]).font = normal_font
        ws_risk.cell(row=i, column=2, value=round(d["GMV"], 2)).font = normal_font
        ws_risk.cell(row=i, column=3, value=d["RMA%"]).font = normal_font
        ws_risk.cell(row=i, column=4, value=d["等级"]).font = normal_font
        ws_risk.cell(row=i, column=5, value=d["健康度评分"]).font = normal_font
    auto_col_width(ws_risk)

    # 散点图
    scatter = ScatterChart()
    scatter.title = "卖家风险矩阵：GMV vs RMA%"
    scatter.x_axis.title = "GMV ($)"
    scatter.y_axis.title = "RMA% (退货率)"
    scatter.style = 10
    scatter.width = 22
    scatter.height = 14
    xvalues = Reference(ws_risk, min_col=2, min_row=2, max_row=total_sellers + 1)
    yvalues = Reference(ws_risk, min_col=3, min_row=2, max_row=total_sellers + 1)
    series = scatter.series
    from openpyxl.chart import Series
    s = Series(yvalues, xvalues, title="卖家")
    scatter.series.append(s)
    ws_risk.add_chart(scatter, "G1")

    # 风险说明
    ws_risk.cell(row=total_sellers + 3, column=1, value="【风险矩阵解读】").font = title_font
    ws_risk.cell(row=total_sellers + 4, column=1, value="右上角（高GMV高RMA%）：优先跟进，解决退货问题").font = normal_font
    ws_risk.cell(row=total_sellers + 5, column=1, value="左上角（低GMV高RMA%）：评估是否清退").font = normal_font
    ws_risk.cell(row=total_sellers + 6, column=1, value="右下角（高GMV低RMA%）：核心卖家，维持现状").font = normal_font
    ws_risk.cell(row=total_sellers + 7, column=1, value="左下角（低GMV低RMA%）：关注成长潜力").font = normal_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), date_range
